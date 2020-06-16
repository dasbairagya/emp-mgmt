#!/usr/bin/env python
# Dev: Gopal Dasbairagya
# 15 Jun, 2020
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import PatternFill
from tkinter import messagebox
from tkinter import ttk #provite combobox
from tkinter import *
import openpyxl
import os
class Employee:
    '''frame is responsible for the design '''
    def __init__(self, root):
        self.root = root
        self.path = 'emp.xlsx'
        self.root.title("Employee Management System")
        self.root.geometry("1350x730+100+70")

        title = Label(self.root, text="Employee Management System", bd=10, relief=GROOVE, font=("times new roman", 30, "bold"), bg="yellow", fg="black")
        title.pack(side=TOP, fill=X)

        #all variables
        self.roll_no = StringVar()
        self.name = StringVar()
        self.email = StringVar()
        self.phone = StringVar()
        self.gender = StringVar()
        self.dob = StringVar()
        self.search_by = StringVar()
        self.search_txt = StringVar()


        #1.Form Frame ----------------------------------------------------------------------------------------------------
        #form frame will place in root frame
        form_frame = Frame(self.root, bd=4, relief=RIDGE, bg="orange")
        form_frame.place(x=20, y=100, width=450, height=600)

        #form title
        form_title = Label(form_frame, text="Manage Employees", bg="orange", fg="white", font=("times new roman", 20, "bold"))
        form_title.grid(row=0, columnspan=2, pady=20)

        #fields & labels will be place in form frame
        #roll
        lbl_roll = Label(form_frame, text="Emp Id", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_roll.grid(row=1, column=0, pady=10, padx=20, sticky='w')

        txt_roll = Entry(form_frame, textvariable=self.roll_no, font=("times new roman", 16, "bold"), bd=5, relief=GROOVE)
        txt_roll.grid(row=1, column=1, pady=10, padx=20, sticky='w')

        #name
        lbl_name = Label(form_frame,  text="Name", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_name.grid(row=2, column=0, pady=10, padx=20, sticky='w')

        txt_name = Entry(form_frame, textvariable=self.name, font=("times new roman", 16, "bold"), bd=5, relief=GROOVE)
        txt_name.grid(row=2, column=1, pady=10, padx=20, sticky='w')
        #email
        lbl_email = Label(form_frame, text="Email", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_email.grid(row=3, column=0, pady=10, padx=20, sticky='w')

        txt_email = Entry(form_frame, textvariable=self.email,  font=("times new roman", 16, "bold"), bd=5, relief=GROOVE)
        txt_email.grid(row=3, column=1, pady=10, padx=20, sticky='w')

        #gender combobox
        lbl_gender = Label(form_frame, text="Gender", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_gender.grid(row=4, column=0, pady=10, padx=20, sticky='w')

        # list = ['Male', 'Female']
        # combo_gender = OptionMenu(form_frame, self.gender, *list)
        # combo_gender.config(width=25)
        # self.gender.set('Male')
        combo_gender = ttk.Combobox(form_frame, textvariable=self.gender,  font=("times new roman", 16, "bold"), state="readonly")
        combo_gender['values']= ('Male', 'Female', 'Other')
        # self.gender.set('Male')
        combo_gender.grid(row=4, column=1, pady=10, padx=20, sticky='w')
        lbl_contact = Label(form_frame, text="Contact", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_contact.grid(row=5, column=0, pady=10, padx=20, sticky='w')

        txt_contact = Entry(form_frame, textvariable=self.phone,  font=("times new roman", 16, "bold"), bd=5, relief=GROOVE)
        txt_contact.grid(row=5, column=1, pady=10, padx=20, sticky='w')
        #dob
        lbl_dob = Label(form_frame, text="D.O.B", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_dob.grid(row=6, column=0, pady=10, padx=20, sticky='w')

        txt_dob = Entry(form_frame, textvariable=self.dob,  font=("times new roman", 16, "bold"), bd=5, relief=GROOVE)
        txt_dob.grid(row=6, column=1, pady=10, padx=20, sticky='w')
        #adress
        lbl_address = Label(form_frame, text="Address", bg="orange", fg="white", font=("times new roman", 16, "bold"))
        lbl_address.grid(row=7, column=0, pady=10, padx=20, sticky='w')

        self.txt_address = Text(form_frame, width=29, height=4)
        self.txt_address.grid(row=7, column=1, pady=10, padx=20, sticky='w')

        #2.Button Frame -------------------------------------------------------------------------------------------------
        # Button frame will place in form frame
        btn_frame = Frame(form_frame, bd=4, relief=RIDGE, bg="orange")
        btn_frame.place(x=20, y=520, width=400)

        #Buttons
        #buttons will place in button frame
        btnAdd = Button(btn_frame, text="Add", width=5, bg='green', bd=4, command=self.add_emp).grid(row=0, column=0, padx=10, pady=10)
        btnUpdate = Button(btn_frame, text="Update", width=5, bg='yellow', bd=4, command=self.update_emp).grid(row=0, column=1, padx=10, pady=10)
        btnDelete = Button(btn_frame, text="Delete", width=5, bg='crimson', bd=4, command=self.delete_emp).grid(row=0, column=2, padx=10, pady=10)
        btnClear = Button(btn_frame, text="Clear", width=5, bg='gray', bd=4, command=self.clear).grid(row=0, column=3, padx=10, pady=10)


        #3.Details Frame -------------------------------------------------------------------------------------------------
        #details frame will place in root frame
        detail_frame = Frame(self.root, bd=4, relief=RIDGE, bg="yellow")
        detail_frame.place(x=500, y=100, width=800, height=600)

        #fields & label will be place in details frame
        #search By
        lbl_search = Label(detail_frame, text="Search By", bg="yellow", fg="black", font=("times new roman", 16, "bold"))
        lbl_search.grid(row=0, column=0, pady=10, padx=15, sticky='w')

        combo_search = ttk.Combobox(detail_frame, textvariable=self.search_by, width=10, font=("times new roman", 16, "bold"), state="readonly")
        combo_search['values'] = ('Roll_no', 'Name', 'contact')
        combo_search.grid(row=0, column=1, pady=10, padx=15, sticky='w')

        #Search
        txt_search = Entry(detail_frame, textvariable=self.search_txt, width=15, font=("times new roman", 16, "bold"), bd=5, relief=GROOVE)
        txt_search.grid(row=0, column=2, pady=10, padx=15, sticky='w')
        #buttons
        btnSearch = Button(detail_frame, text="Search", width=5, bd=4, pady=5, command=self.search_data).grid(row=0, column=3, padx=10, pady=10)
        btnShow = Button(detail_frame, text="Show All", width=5, bd=4, pady=5, command=self.fetch_data).grid(row=0, column=4, padx=10, pady=10)
        btnShow = Button(detail_frame, text="Exit", width=5, bd=4, pady=5, command=self.exit_app).grid(row=0, column=5, padx=10, pady=10)

        #4.Table Frame -------------------------------------------------------------------------------------------------
        # table frame will place in details frame
        table_frame = Frame(detail_frame, bd=4, relief=RIDGE, bg="crimson")
        table_frame.place(x=10, y=70, width=760, height=510)

        scroll_x = Scrollbar(table_frame, orient=HORIZONTAL)
        scroll_y = Scrollbar(table_frame, orient=VERTICAL)
        self.data_table = ttk.Treeview(table_frame, column=("roll", "name", "email", "gender", "phone", "dob", "address"), xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=self.data_table.xview)
        scroll_y.config(command=self.data_table.yview)
        self.data_table.heading('roll', text="Id")
        self.data_table.heading('name', text="Name")
        self.data_table.heading('email', text="Email")
        self.data_table.heading('phone', text="Contact")
        self.data_table.heading('gender', text="Gender")
        self.data_table.heading('dob', text="DOB")
        self.data_table.heading('address', text="Address")
        self.data_table['show']='headings'
        self.data_table.column("roll", width=50)
        self.data_table.column("name", width=100)
        self.data_table.column("email", width=150)
        self.data_table.column("gender", width=80)
        self.data_table.column("phone", width=100)
        self.data_table.column("dob", width=100)
        self.data_table.column("address", width=150)
        self.data_table.pack(fill=BOTH, expand=1)
        self.data_table.bind("<ButtonRelease-1>", self.get_cursor)
        self.data_table.tag_configure('odd', background='#E8E8E8')
        self.data_table.tag_configure('even', background='#DFDFDF')
        self.fetch_data()

    def add_emp(self):
        if self.roll_no.get()!="" and self.name.get()!="" and self.email.get()!="" and self.gender.get()!="":

            if os.path.exists(self.path): # if file exits load the file

                book = openpyxl.load_workbook(self.path)
                # my_sheet = my_wb_obj.['daily sale'] to open a specific xl book
                sheet = book.active
                # get the max row and max column
                max_row = sheet.max_row
                max_col = sheet.max_column

                # get input values
                id = self.roll_no.get()
                name = self.name.get()
                email = self.email.get()
                gender = self.gender.get()
                contact = self.phone.get()
                dob = self.dob.get()
                address = self.txt_address.get('1.0', END)
                datas = [id, name, email, gender, contact, dob, address]

                # alert if duplicate entry
                result = False
                for row in range(2, max_row + 1):
                    # print(my_sheet_obj.cell(row=row, column=1).value)
                    if sheet.cell(row=row, column=1).value == id:
                        result = True
                if result == True:
                    messagebox.showerror("Error", "Duplicate Entry!!!")
                    return
                else:
                    # store the input values in a single row after max row no.
                    i = 1
                    for data in datas:
                        # print(data)
                        cell = sheet.cell(max_row + 1, column=i)
                        cell.value = data
                        i += 1
                    book.save(self.path)
                    self.fetch_data()
                    self.clear()
                    messagebox.showinfo("Success", "Record added successfully!")


            else: # if file does not exits create the file
                my_wb = openpyxl.Workbook()
                my_sheet = my_wb.active
                # add header
                headers = ['ID', 'NAME', 'EMAIL', 'GENDER', 'CONTACT', 'DOB', 'ADDRESS']
                col = 1
                for header in headers:
                    cell = my_sheet.cell(row=1, column=col)
                    cell.value = header
                    col += 1

                # get input values
                id = self.roll_no.get()
                name = self.name.get()
                email = self.email.get()
                gender = self.gender.get()
                contact = self.phone.get()
                dob = self.dob.get()
                address = self.txt_address.get('1.0', END)
                datas = [id, name, email, gender, contact, dob, address]

                col = 1
                for data in datas:
                    cell = my_sheet.cell(row=2, column=col)
                    cell.value = data
                    col += 1

                # save the file thereafter
                my_wb.save(self.path)
                self._format_header_cell()
                self.fetch_data()
                self.clear()
                messagebox.showinfo("Success", "Record has been created!")

        else:
            messagebox.showerror("Error", "All fields are required!!!")

    def _format_header_cell(self):
        book = openpyxl.load_workbook(self.path)
        sheet = book.active
        max_col = sheet.max_column
        for col in range(1, max_col + 1):
            cell_title = sheet.cell(1, col)
            cell_title.fill = PatternFill(start_color="8a2be2", end_color="8a2be2", fill_type="solid")
        book.save(self.path)

    def fetch_data(self):
        if os.path.exists(self.path):
            book = openpyxl.load_workbook(self.path)
            sheet = book.active
            max_row = sheet.max_row
            max_col = sheet.max_column

            # print all rows value
            if max_row > 1:
                self.data_table.delete(*self.data_table.get_children())
                # print all rows value
                for row in range(2, max_row + 1):
                    mylist = []
                    for column in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=row, column=column)
                        mylist.append(cell_obj.value)
                    mytuple = tuple(mylist)

                    if row % 2 == 0:
                        self.data_table.insert('', END, values=mytuple,  tags="even")
                    else:
                        self.data_table.insert('', END, values=mytuple,  tags="odd")
                self.search_by.set("")
                self.search_txt.set("")
        else:
            messagebox.showinfo("Info", "No record created yet!")

    def clear(self):
        self.roll_no.set("")
        self.name.set("")
        self.email.set("")
        self.phone.set("")
        self.gender.set("")
        self.dob.set("")
        self.txt_address.delete("1.0", END)

    def get_cursor(self, ev):
        try:
            cursor_row = self.data_table.focus()
            if cursor_row:
                content = self.data_table.item(cursor_row)
                row = content['values']
                # print(row)
                self.roll_no.set(row[0])
                self.name.set(row[1])
                self.email.set(row[2])
                self.gender.set(row[3])
                self.phone.set(row[4])
                self.dob.set(row[5])
                self.txt_address.delete("1.0", END)
                self.txt_address.insert(END, row[6])
            else:
                pass
        except:
            pass


    def update_emp(self):
        if self.roll_no.get() != "":
            # get input values
            id = self.roll_no.get()
            name = self.name.get()
            email = self.email.get()
            gender = self.gender.get()
            contact = self.phone.get()
            dob = self.dob.get()
            address = self.txt_address.get('1.0', END)
            datas = [name, email, gender, contact, dob, address]

            my_wb_obj = openpyxl.load_workbook(self.path)
            my_sheet_obj = my_wb_obj.active
            # get the max row and max column
            max_row = my_sheet_obj.max_row
            max_col = my_sheet_obj.max_column
            result = False
            matched_row = None
            for row in range(2, max_row + 1):
                # print(my_sheet_obj.cell(row=row, column=1).value)
                if my_sheet_obj.cell(row=row, column=1).value == id:
                    result = True
                    matched_row = row

            if result == True:  # alert if id does not exit else update success
                # print(matched_row)
                # store the input values in a matched row found.
                i = 2
                for data in datas:
                    # print(data)
                    cell = my_sheet_obj.cell(matched_row, column=i)
                    cell.value = data
                    i += 1
                my_wb_obj.save(self.path)
                messagebox.showinfo('Updated', 'Employee has been updated successfully!', icon='info')
                self.fetch_data()
                self.clear()

            else:
                messagebox.showerror("Error", "Record not found!")
        else:
            messagebox.showerror("Error", "Please select an employee!")


    def delete_emp(self):
        if self.roll_no.get()!="":
            id = self.roll_no.get()
            response = self.alert()
            if response=='yes':
                book = openpyxl.load_workbook(self.path)
                sheet = book.active
                # get the max row and max column
                max_row = sheet.max_row
                max_col = sheet.max_column
                result = False
                matched_row = None
                for row in range(2, max_row + 1):
                    # print(my_sheet_obj.cell(row=row, column=1).value)
                    if sheet.cell(row=row, column=1).value == id:
                        result = True
                        matched_row = row

                if result == True:
                    print(f"Deleted row : {matched_row}")
                    sheet.delete_rows(matched_row)
                    book.save(self.path)
                    self.fetch_data()
                    self.clear()

                else:
                    messagebox.showerror("Error", "Record not found!")

            else:
                pass
        else:
            messagebox.showerror("Error", "Please select an employee!")


    def exit_app(self):
        MsgBox = messagebox.askquestion('Exit Application', 'Are you sure you want to exit the application',
                                           icon='question')
        if MsgBox == 'yes':
            root.destroy()
        else:
            pass
            # messagebox.showinfo('Return', 'You will now return to the application screen')

    def alert(self):
        MsgBox = messagebox.askquestion('Delete', 'Are you sure you want to delete the record',
                                        icon='warning')
        return MsgBox

    def search_data(self):
        if self.search_by.get()=="":
            messagebox.showerror("Error", "Provide search by value")
        elif self.search_txt.get()=="":
            messagebox.showerror("Error", "Provide search keyword")
        else:
            if not os.path.exists(self.path):
                messagebox.showinfo("Info", "No record created yet!")
                return
            search_by = self.search_by.get()
            search_key_word = self.search_txt.get()
            book = openpyxl.load_workbook(self.path)
            sheet = book.active
            # get the max row and max column
            max_row = sheet.max_row
            max_col = sheet.max_column
            result = False
            names = []
            #get the list of cell values of perticular column here 2 i.e NAMES and make a list
            for row in sheet.iter_rows(2, max_row):
                names.append(row[1].value)
            # print(names)

            #get the list of matching names by search_key_word
            matching = [s for s in names if search_key_word in s]
            # print(matching)


            #todo: make the list of corresponding rows from matching names

            # print("-----------------------")
            #todo: getting the row and numbers from coordinate value in openpyxl

            matched_cell = []  #['<Cell 'Sheet'.B5>', 'B6']
            matched_rows = []
            for item in matching:
                for row in sheet.iter_rows(2, max_row):
                    # print(row[1].coordinate) get the coordinate value e.g. B5 from the cell object
                    if row[1].value == item:
                        matched_cell.append(row[1])
                        xy = coordinate_from_string(row[1].coordinate)  # returns tuple like ('B',5)
                        # print(xy)
                        row = xy[1]
                        matched_rows.append(row)

            # print(matched_cell)
            # print(matched_rows)

            #todo: loop through the rows and make data tuple for every row
            datalist = []
            for row in matched_rows:
                datavalues = []
                for col in range(1, max_col):
                    # get the all cell value of matched row and make tuples
                    datavalues.append(sheet.cell(row=row, column=col).value)

                data_tuple = tuple(datavalues)
                # print(data_tuple)
                datalist.append(data_tuple)
                tuple_of_tuples = tuple(datalist)
                # print(tuple_of_tuples)

                if len(tuple_of_tuples) > 0:
                    self.data_table.delete(*self.data_table.get_children())
                    count=0
                    for row in tuple_of_tuples:
                        count+=1
                        if count % 2 == 0:
                            self.data_table.insert('', END, values=row, tags="even")  # tags used for odd-even row color
                        else:
                            self.data_table.insert('', END, values=row, tags="odd")

                else:
                   messagebox.showerror("Not Found", "No record found!", icon='warning')





if __name__ == '__main__':
    root = Tk()
    ob = Employee(root)
    root.mainloop()
